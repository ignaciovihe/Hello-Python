### File Handling ###

import xml
import csv
import json
import os

script_dir = os.path.dirname(os.path.abspath(__file__))

# .txt file

# Leer, escribir y sobrescribir si ya existe
txt_file = open(f"{script_dir}/my_file.txt", "w+")

txt_file.write(
    "Mi nombre es Brais\nMi apellido es Moure\n35 años\nY mi lenguaje preferido es Python")

# Posiciona el cursor al inicio del fichero
txt_file.seek(0)

# Lee e imprime todo el contenido del fichero
print(txt_file.read())

# Lee e imprime 10 caracteres desde el inicio del fichero
txt_file.seek(0)
print(txt_file.read(10))

# Lee e imprime el resto de la línea actual desde la posición 11
print(txt_file.readline())

# Lee e imprime la siguiente línea
print(txt_file.readline())

# Lee e imprime las líneas restantes del fichero
for line in txt_file.readlines():
    print(line)

# Escribe una nueva línea en el fichero
txt_file.write("\nAunque también me gusta Kotlin")

# Posiciona el cursor al inicio del fichero, lee e imprime todo su contenido
txt_file.seek(0)
print(txt_file.read())

# Cierra el fichero
txt_file.close()

# Agrega una nueva línea en el fichero
with open(f"{script_dir}/my_file.txt", "a") as my_other_file:
    my_other_file.write("\nY Swift")

# os.remove("Intermediate/my_file.txt")

# .json file

json_file = open(f"{script_dir}/my_file.json", "w+")

json_test = {
    "name": "Brais",
    "surname": "Moure",
    "age": 35,
    "languages": ["Python", "Swift", "Kotlin"],
    "website": "https://moure.dev"}

json.dump(json_test, json_file, indent=2)

json_file.close()

with open(f"{script_dir}/my_file.json") as my_other_file:
    for line in my_other_file.readlines():
        print(line)

json_dict = json.load(open(f"{script_dir}/my_file.json"))
print(json_dict)
print(type(json_dict))
print(json_dict["name"])

# .csv file


csv_file = open(f"{script_dir}/my_file.csv", "w+")

csv_writer = csv.writer(csv_file)
csv_writer.writerow(["name", "surname", "age", "language", "website"])
csv_writer.writerow(["Brais", "Moure", 35, "Python", "https://moure.dev"])
csv_writer.writerow(["Roswell", "", 2, "COBOL", ""])

csv_file.close()

with open(f"{script_dir}/my_file.csv") as my_other_file:
    for line in my_other_file.readlines():
        print(line)

# .xlsx file
# import xlrd # Debe instalarse el módulo
# Para xlsx modernos se usa openpyxl

from openpyxl import Workbook, load_workbook 

import os

file_path = os.path.join(script_dir, "datos.xlsx")

# 0. Si el archivo no existe, creamos uno nuevo
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    ws["A1"] = "Nombre"
    ws["B1"] = "Edad"
    ws.append(["Ana", 25])
    ws.append(["Luis", 30])
    wb.save(file_path)
    print(f"Archivo creado: {file_path}")

# 1. Abrir el archivo Excel (similar a open_workbook de xlrd)
workbook = load_workbook(f"{script_dir}/datos.xlsx")

# 2. Seleccionar una hoja
sheet = workbook.active          # primera hoja activa
# sheet = workbook["Hoja1"]      # o por nombre

# 3. Leer valores de celdas
print(sheet["A1"].value)        # celda A1
print(sheet.cell(row=1, column=1).value)  # fila 1, columna 1

# 4. Recorrer filas (como readlines() en archivos de texto)
for row in sheet.iter_rows(values_only=True):
    print(row)

# .xml file

# ¿Te atreves a practicar cómo trabajar con este tipo de ficheros?

import xml.etree.ElementTree as ET

# 0. Carpeta del script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Ruta del archivo XML
file_path = os.path.join(script_dir, "datos.xml")

# 1. Si el archivo no existe, lo creamos
if not os.path.exists(file_path):
    root = ET.Element("Personas")  # raíz

    persona1 = ET.SubElement(root, "Persona")
    ET.SubElement(persona1, "Nombre").text = "Ana"
    ET.SubElement(persona1, "Edad").text = "25"

    persona2 = ET.SubElement(root, "Persona")
    ET.SubElement(persona2, "Nombre").text = "Luis"
    ET.SubElement(persona2, "Edad").text = "30"

    tree = ET.ElementTree(root)
    tree.write(file_path, encoding="utf-8", xml_declaration=True)
    print(f"Archivo XML creado en {file_path}")

# 2. Abrir el XML existente
tree = ET.parse(file_path)
root = tree.getroot()

# 3. Leer valores de nodos
for persona in root.findall("Persona"):
    nombre = persona.find("Nombre").text
    edad = persona.find("Edad").text
    print(f"{nombre} - {edad}")

# 4. También puedes acceder directamente a un nodo específico
primera_persona = root.find("Persona")
print("Primera persona:", primera_persona.find("Nombre").text)
